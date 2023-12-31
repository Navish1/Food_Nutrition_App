import streamlit as st
from PIL import Image
from tensorflow.keras.utils import load_img, img_to_array
import numpy as np
from keras.models import load_model
import pandas as pd
from openpyxl import load_workbook
from PIL import Image

st.set_page_config(page_title="FoodiesTrip", page_icon=":🍔:", layout="wide")


df = pd.read_excel(
                    io="food_nutrient.xlsx",
                    engine='openpyxl',
                    usecols="A:H",)

df1 = df.astype(str)


wb = load_workbook("food_nutrient.xlsx")
sheet = wb.active

model = load_model("inceptionv3_sgd.hdf5")

labels = {0: 'apple_pie',
            1: 'baby_back_ribs',
            2: 'baklava',
            3: 'beef_carpaccio',
            4: 'beef_tartare',
            5: 'beet_salad',
            6: 'beignets',
            7: 'bibimbap',
            8: 'bread_pudding',
            9: 'breakfast_burrito',
            10:'bruschetta',
            11:'caesar_salad',
            12: 'cannoli',
            13: 'caprese_salad',
            14: 'carrot_cake',
            15: 'ceviche',
            16: 'cheese_plate',
            17: 'cheesecake',
            18: 'chicken_curry',
            19: 'chicken_quesadilla',
            20: 'chicken_wings',
            21: 'chocolate_cake',
            22: 'chocolate_mousse',
            23: 'churros',
            24: 'clam_chowder',
            25: 'club_sandwich',
            26: 'crab_cakes',
            27: 'creme_brulee',
            28: 'croque_madame',
            29: 'cup_cakes',
            30: 'deviled_eggs',
            31: 'donuts',
            32: 'dumplings',
            33: 'edamame',
            34: 'eggs_benedict',
            35: 'escargots',
            36: 'falafel',
            37: 'filet_mignon',
            38: 'fish_and_chips',
            39: 'foie_gras',
            40: 'french_fries',
            41: 'french_onion_soup',
            42: 'french_toast',
            43: 'fried_calamari',
            44: 'fried_rice',
            45: 'frozen_yogurt',
            46: 'garlic_bread',
            47: 'gnocchi',
            48: 'greek_salad',
            49: 'grilled_cheese_sandwich',
            50: 'grilled_salmon',
            51: 'guacamole',
            52: 'gyoza',
            53: 'hamburger',
            54: 'hot_and_sour_soup',
            55: 'hot_dog',
            56: 'huevos_rancheros',
            57: 'hummus',
            58: 'ice_cream',
            59: 'lasagna',
            60: 'lobster_bisque',
            61: 'lobster_roll_sandwich',
            62: 'macaroni_and_cheese',
            63: 'macarons',
            64: 'miso_soup',
            65: 'mussels',
            66: 'nachos',
            67: 'omelette',
            68: 'onion_rings',
            69: 'oysters',
            70: 'pad_thai',
            71: 'paella',
            72: 'pancakes',
            73: 'panna_cotta',
            74: 'peking_duck',
            75: 'pho',
            76: 'pizza',
            77: 'pork_chop',
            78: 'poutine',
            79: 'prime_rib',
            80: 'pulled_pork_sandwich',
            81: 'ramen',
            82: 'ravioli',
            83: 'red_velvet_cake',
            84: 'risotto',
            85: 'samosa',
            86: 'sashimi',
            87: 'scallops',
            88: 'seaweed_salad',
            89: 'shrimp_and_grits',
            90: 'spaghetti_bolognese',
            91: 'spaghetti_carbonara',
            92: 'spring_rolls',
            93: 'steak',
            94: 'strawberry_shortcake',
            95: 'sushi',
            96: 'tacos',
            97: 'takoyaki',
            98: 'tiramisu',
            99: 'tuna_tartare',
            100: 'waffles'}



def processed_img(img_path):
    img = load_img(img_path, target_size=(299, 299, 3))
    img = img_to_array(img)
    img = img / 255
    img = np.expand_dims(img, [0])
    answer = model.predict(img)
    y_class = answer.argmax(axis=-1)
    print(y_class)
    y = " ".join(str(x) for x in y_class)
    y = int(y)
    res = labels[y]
    print(res)
    return res.capitalize()

def find_row_number(cell_value):
    
    sheet = wb['food_nutrient']

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == cell_value:
                return cell.row

    return None  # Cell value not found

def find_calories(row_cal):
    calories = sheet.cell(row=row_cal,column=4)
    return calories.value

def find_fats(row_cal):
    fats = sheet.cell(row=row_cal,column=5)
    return fats.value

def find_carbs(row_cal):
    carbohydrate = sheet.cell(row=row_cal,column=6)
    return carbohydrate.value

def find_fiber(row_cal):
    fiber = sheet.cell(row=row_cal,column=7)
    return fiber.value

def find_protein(row_cal):
    protein = sheet.cell(row=row_cal,column=8)
    return protein.value

def run():
    
    st.sidebar.title("FoodiesTrip")
    st.sidebar.write("""
    FoodiesTrip is an end-to-end **CNN Image Classification Model** which identifies the food in your image and provide its nutritional content. 

    It can identify over 100 different food classes

    It is based upon a pre-trained Image Classification Model that comes with Keras and then retrained on the infamous **Food101 Dataset**.

    **Accuracy :** **`81.26%`**

    **Model :** **`InceptionV3`**

    **Dataset :** **`Food101`**
    \n\n
    
    The Food categories available are:\n
    • apple pie\n
    • baby back ribs\n
    • baklava\n
    • beef carpaccio\n
    • beef tartare\n
    • beet salad\n
    • beignets\n
    • bibimbap\n
    • bread pudding\n
    • breakfast burrito\n
    • bruschetta\n
    • caesar salad\n
    • cannoli\n
    • caprese salad\n
    • carrot cake\n
    • ceviche\n
    • cheesecake\n
    • cheese plate\n
    • chicken curry\n
    • chicken quesadilla\n
    • chicken wings\n
    • chocolate cake\n
    • chocolate mousse\n
    • churros\n
    • clam chowder\n
    • club sandwich\n
    • crab cakes\n
    • creme brulee\n
    • croque madame\n
    • cup cakes\n
    • deviled eggs\n
    • donuts\n
    • dumplings\n
    • edamame\n
    • eggs benedict\n
    • escargots\n
    • falafel\n
    • filet mignon\n
    • fish and chips\n
    • foie gras\n
    • french fries\n
    • french onion_soup\n
    • french toast\n
    • fried calamari\n
    • fried rice\n
    • frozen yogurt\n
    • garlic bread\n
    • gnocchi\n
    • greek salad\n
    • grilled cheese sandwich\n
    • grilled salmon\n
    • guacamole\n
    • gyoza\n
    • hamburger\n
    • hot and sour soup\n
    • hot dog\n
    • huevos rancheros\n
    • hummus\n
    • ice cream\n
    • lasagna\n
    • lobster bisque\n
    • lobster roll sandwich\n
    • macaroni and cheese\n
    • macarons\n
    • miso soup\n
    • mussels\n
    • nachos\n
    • omelette\n
    • onion rings\n
    • oysters\n
    • pad thai\n
    • paella\n
    • pancakes\n
    • panna cotta\n
    • peking duck\n
    • pho\n
    • pizza\n
    • pork chop\n
    • poutine\n
    • prime rib\n
    • pulled pork sandwich\n
    • ramen\n
    • ravioli\n
    • red velvet cake\n
    • risotto\n
    • samosa\n
    • sashimi\n
    • scallops\n
    • seaweed salad\n
    • shrimp and grits\n
    • spaghetti bolognese\n
    • spaghetti carbonara\n
    • spring rolls\n
    • steak\n
    • strawberry shortcake\n
    • sushi\n
    • tacos\n
    • takoyaki\n
    • tiramisu\n
    • tuna tartare\n
    • waffles\n
    """)
    
    st.title("Food Recognition and Nutrition Estimation 🍅")
    # img_file = st.file_uploader("Choose an Image", type=["jpg", "png"])
    # img_file = st.camera_input(label = "Take a pic of your food")

    st.markdown("---")

    choice = st.radio("Select an option to import your food:", ("Upload Image", "Take a Picture"))

    img_file = None  # Define the variable outside the if blocks

    st.markdown("---")

    if choice == "Upload Image":
        img_file = st.file_uploader("Choose an Image", type=["jpg", "png"])
        # if img_file is not None:
        #     process_image(img_file)

    elif choice == "Take a Picture":
        img_file = st.camera_input(label="Take a pic of your food")
        # if img is not None:
        #     process_image(img)

    st.markdown("---")
    if img_file is not None:
        img = Image.open(img_file).resize((299, 299))
        st.markdown("---")
        st.image(img, use_column_width=False)
        save_image_path = 'upload_images/' + img_file.name
        with open(save_image_path, "wb") as f:
            f.write(img_file.getbuffer())

        # if st.button("Predict"):
        if img_file is not None:
            result = processed_img(save_image_path)
            print(result)
            st.success("**Predicted : " + result + '**')
            st.markdown("---")
            result_id = str(result)
            
            row_number = find_row_number(result_id)

            cal = str(find_calories(row_number))
            pro = str(find_protein(row_number))
            fat = str(find_fats(row_number))
            carbs = str(find_carbs(row_number))
            fib = str(find_fiber(row_number))


            st.error("Servings per 100 g")
            
            st.warning('**' + "Calories : " + cal + ' g' '**')
            
            st.warning('**' + "Protein : " + pro + ' g' + '**')
            
            st.warning('**' + "Fats : " + fat + ' g' + '**')
            
            st.warning('**' + "Carbohydrates : " + carbs + ' g' + '**')
            
            st.warning('**' + "Fibers : " + fib + ' g' + '**')
            

    # st.dataframe(df1,use_container_width=True) 


run()
